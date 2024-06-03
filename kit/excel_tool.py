import openpyxl
from tqdm import tqdm
from kit.pytools import create_folder, write_txt, write_key, convert_keys, vocab_append

def check_header(sh):
    if (sh.cell(row = 1, column=1).value == 'ID' and sh.cell(row = 1, column=2).value == 'Text' and sh.cell(row = 1, column=3).value == 'Key'): 
        return True
    else: 
        return False

def read_excel(filepath: str):
    
    try:  
        wb = openpyxl.load_workbook(filepath)
        sh = wb.active
        m_row = sh.max_row
        vocab_array = []
        if not check_header(sh):
            print('Errore: intestazione file CSV non conforme')
            raise
        else:
            foldername = create_folder(filepath)
            for i in tqdm(range(2, m_row+1), total=m_row+1, desc='Conversione .txt/.key:', bar_format='{l_bar}{bar} {n_fmt}/{total_fmt}'):
                
                filename = sh.cell(row = i, column=1).value
                text = sh.cell(row = i, column=2).value
                keys = convert_keys(sh.cell(row = i, column=3).value)

                txt_file = foldername + '/' + filename + '.txt'
                if write_txt(txt_file, text): pass
                else: 
                    print('Errore: la riga ' + str(i) + '  contiene un valore non valido')
                    continue
                
                key_file = foldername + "/" + filename + ".key"
                if write_key(key_file, keys): pass
                else: print('Errore durante la scrittura di ' + filename + '.key')

                vocab_array = vocab_append(vocab_array, keys) 
            
            return vocab_array
    
    except:
        raise Exception()
    
def vocab_excel(filepath: str):
    try:  
        vocab_array = []
        wb = openpyxl.load_workbook(filepath)
        sh = wb.active
        m_row = sh.max_row
        if not check_header(sh):
            print('Errore: intestazione file CSV non conforme')
            raise
        else:
            for i in range(2, m_row + 1):
                keys = convert_keys(sh.cell(row = i, column=3).value)
                vocab_array = vocab_append(vocab_array, keys)
            return vocab_array
    
    except:
        raise Exception()